from re import I
import data_grab as dg
import LSTM as L
import buy_sell as bs
import pandas as pd
from openpyxl import workbook #pip install openpyxl
from openpyxl import load_workbook
#import LSTM_optimal as LO


#could make it so we input a list of stocks/strategies and loops through the list
def main():
    #table=pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')
    #names = list(table[0]['Symbol'])
    #names = []
    #names.append('AMD')
    #pl = {}
    #strats = []
    #strats.append('darse')
    #strats.append('full')
    #strats.append('dense')
    #i = 0
    # for stock in names:
    #     data = dg.get_data(stock)
    #     for strat in strats:
    #         listy = []
    #         y_pred, y_real = L.lstm_model(strat, data, 1, stock)
    #         performance, corr, up, down = L.perf_metrics(y_pred, y_real, data, stock)
    #         L.graph_LSTM_results(y_real, y_pred, stock, strat)
    #         money = bs.compare_portfolios(y_real, y_pred, 10000)
    #         listy.append(money['buy and hold'].iloc[0])
    #         listy.append(money['trader value'].iloc[0])
    #         key = str(stock)+str(strat)
    #         pl[key] = listy
    #     i = i + 1
    # print(pl)
    # j = 0
    # buy_hold = 0
    # darse = 0
    # full = 0
    # dense = 0
    # for key in pl:
    #     #darse
    #     if j % 3 == 0:
    #         if pl[key][0] > pl[key][1]:
    #             buy_hold = buy_hold + 1
    #         else:
    #             darse = darse + 1
    #     #full
    #     if j % 3 == 1:
    #         if pl[key][0] > pl[key][1]:
    #             buy_hold = buy_hold + 1
    #         else:
    #             full = full + 1
    #     #dense
    #     if j % 3 == 2:
    #         if pl[key][0] > pl[key][1]:
    #             buy_hold = buy_hold + 1
    #         else:
    #             dense = dense + 1
    #     j = j + 1
    # print(j)
    # print(buy_hold)
    # print(darse)
    # print(full)
    # print(dense)






    #data = dg.get_data('SPY')
    #for strat in strats:
    #listy = []
    ####if using full then days ahead must be 1?
    #y_pred, y_real = L.lstm_model('dense', data, 3, 'SPY')
    #performance, corr, up, down = L.perf_metrics(y_pred, y_real, data, 'SPY')
    #L.graph_LSTM_results(y_real, y_pred, 'SPY', 'darse')
    #money = bs.compare_portfolios(y_real, y_pred, 10000)
    #listy.append(money['buy and hold'].iloc[0])
    #listy.append(money['trader value'].iloc[0])
    #key = str(stock)+str(strat)
    #pl[key] = listy
    # print(corr)
    # print(up)
    # print(down)
    # print(performance)
    # print(money)



    #WITH OPTIMIZATUION USING KERAS TUNER#
    #stock = 'GME'
    #strategy = 'darse'
    #data = dg.get_data(stock)
    strats = []
    strats.append('darse')
    #strats.append('full')
    #strats.append('dense')
    #strats.append('close')
    #strats.append('all feat difference results')
    table=pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')
    names = list(table[0]['Symbol'])
    wb = load_workbook("/mnt/c/Users/Ray/Desktop/4_1_preset_model.xlsx")
    ws = wb.active
    for stock in names[:1]:
        data = dg.get_data(stock)
        i=2
        open_row = ws.max_row+1
        for strategy in strats:
            y_pred, y_real, history = L.lstm_model(str(strategy), data, 1, stock, 224, 256, 256, 224, .0006, 'tanh', 'tanh', 'tanh', 'tanh', .3, .3, .25, .1)
            performance, corr, up, down, tp, fp, tn, fn, total, f1, f2 = L.perf_metrics(y_pred, y_real, data, stock)
            L.graph_LSTM_results(y_real, y_pred, stock, str(strategy))
            L.graph_val(stock, strategy, history)
            #money = bs.compare_portfolios(y_real, y_pred, 10000)
            L.pretty_printer(corr, up, down, performance, tp, fp, tn, fn, total, f1, f2)
            #write results to sheet
            #wb = load_workbook("/mnt/c/Users/Ray/Desktop/4_1_preset_model.xlsx")
            #ws = wb.active
            string = str(round(performance.iloc[0][0], 2)) #+ '/' + str(round(performance.iloc[0][1], 2)) + '/' + str(round(performance.iloc[0][2], 2))
            #open_row = ws.max_row+1
            ws.cell(column=1, row=open_row, value=str(stock))
            ws.cell(column=i, row=open_row, value=string)
            #wb.save("/mnt/c/Users/Ray/Desktop/4_1_preset_model.xlsx")
            i = i + 1
    wb.save("/mnt/c/Users/Ray/Desktop/4_1_preset_model.xlsx")



    """EZ KEY FOR STRATS
    darse
    full
    dense
    close
    all feat difference results
    f"""

if __name__ == "__main__":  
    main()